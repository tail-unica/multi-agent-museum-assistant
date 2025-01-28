import os
import json
from openpyxl import load_workbook
from PIL import Image
#from openpyxl.drawing.image import Image

def parse_excel_to_json(file_path, output_folder="../dummy-data-json_v2", image_folder="../images"):
    # Create output folders
    os.makedirs(output_folder, exist_ok=True)
    os.makedirs(image_folder, exist_ok=True)

    fname = file_path.split("/")[-1]
    # Load the Excel workbook
    workbook = load_workbook(file_path, data_only=True)
    # Prepare JSON structure
    json_data = {}
    image_map = {}
    entry = 0
    for sh_idx, sheet_name in enumerate(workbook.sheetnames):
        #sheet = workbook.active  # Assuming data is in the active sheet
        sheet = workbook[sheet_name]
        # Get headers from the first row
        headers = [cell.value for cell in sheet[1]]

        image_map[sh_idx] = {}

        for image in sheet._images:  # Access embedded images
            # Get the top-left anchor cell of the image
            anchor_cell = image.anchor._from
            row_index = anchor_cell.row
            # Save the image with a unique filename
            image_path = os.path.join(image_folder, f"image_row_{fname}_{sh_idx}_{row_index}.png")

            pil_image = Image.open(image.ref)
            pil_image.save(image_path)
            # Map the image path to the corresponding row
            image_map[sh_idx][row_index] = image_path



        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=False)):
            #row = row[1:]# Skip the header row
            if row_idx != 0:
                row_data = {}
                for header, cell in zip(headers, row):
                    if cell.value:
                        row_data[header] = cell.value

                    # Add the image path if an image is mapped to this row
                    row_data["foto"] = image_map[sh_idx].get(row_idx, None)
                    row_data["room"] = sheet_name

                # Add row data to JSON
                json_data[entry] = row_data
                entry += 1

    # Save JSON output
    file_out = file_path.replace(".xlsx", ".json").split("/")[-1]
    output_file = os.path.join(output_folder, file_out)
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=4)

    print(f"Data successfully parsed and saved to {output_file}")
    return json_data

# File paths
excel_file_path = "/media/Workspace/Projects/multi-agent-museum-assistant/dummy-data-xlsx/TestiGeneraliAllai_20241021.xlsx"
parse_excel_to_json(excel_file_path)
